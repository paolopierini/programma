##### Version 1
from imapclient import IMAPClient
from datetime  import date
import openpyxl

#librerie
import imaplib
import getpass
import pprint
import email

def indirizzo( par ):
        if isinstance(par,tuple):
                if len(par) > 0:
                        if len(par[0]) > 3 and par[0][2] is not None and par[0][3] is not None:
                                return par[0][2].decode()+'@'+par[0][3].decode()
                        else:
                                return 'dati!'
                else:
                        return 'short!'
        else:
                return 'non-tuple!'

imaplib._MAXLINE=10000000

print('Connecting to email account')
server=IMAPClient('mail.esss.lu.se', ssl=True, use_uid=True)
username = 'paolopierini'
p=getpass.getpass()
print('Now I login')
server.login(username,p)
folders=server.list_folders()

for f in folders:
        print(f[2])

folder=server.select_folder('INBOX', readonly=True)
print('{:d} Messages in INBOX'.format(folder[b'EXISTS']))

years=[2017,2018,2019,2020]

# Create the structure
messages={}
for y in years:
        messages[y]=server.search([u'SINCE',date(y,1,1),u'BEFORE',date(y+1,1,1)])
# Print number of msgs
for y,mesgs in messages.items():
        print('{:d} Messages in {:d}'.format(len(mesgs),y))
# Load all messages
allinfo={}
for y,mesgs in messages.items():
        print('Processing {:d}...'.format(y),end='',flush=True)
        n=int(len(mesgs) / 1000)
        allinfo[y]={}
        for i in range(n):
                for k,d in server.fetch(mesgs[1000*i:1000*(i+1)], ['ENVELOPE']).items():
                        allinfo[y][k]=d
                print('{:d}k/'.format(i+1),end='',flush='True')
        for k,d in server.fetch(mesgs[1000*n:], ['ENVELOPE']).items():
                allinfo[y][k]=d
        print('retrieved {:d} messages'.format(len(allinfo[y])),flush=True)

senders={}
wb=openpyxl.Workbook()
for y,mesgs in allinfo.items():
        ws=wb.create_sheet(title=str(y))
        ws.cell(1,1).value='Weekday'
        ws.cell(1,2).value='Year'
        ws.cell(1,3).value='Month'
        ws.cell(1,4).value='Day'
        ws.cell(1,5).value='Time'
        ws.cell(1,6).value='Day of Year'
        ws.cell(1,7).value='From'
        ws.cell(1,8).value='Subject'
        ws.cell(1,9).value="Sent/Received"
        ws.cell(1,10).value="ID"
        senders[y]={}
        i=1
        for id,data in mesgs.items():
                i+=1
                env=data[b'ENVELOPE']
                who=indirizzo(env.from_)
                if who not in senders[y]:
                        print('{:s} sent me a mail!'.format(who))
                        senders[y][who]=1
                else:
                        senders[y][who]+=1
                if env.subject is None:
                        soggetto='N/A'
                else:
                        soggetto=env.subject.decode()
                #print('ID{:d}: From {:s} {:s} received {:}'.format(id,who,soggetto,env.date))
                ws.cell(i,1).value=env.date.isoweekday()
                ws.cell(i,2).value=env.date.year
                ws.cell(i,3).value=env.date.month
                ws.cell(i,4).value=env.date.day
                ws.cell(i,5).value=str(env.date.time())
                ws.cell(i,6).value=env.date.timetuple().tm_yday
                ws.cell(i,7).value=str(who)
                ws.cell(i,8).value=str(soggetto)
                ws.cell(i,9).value='N/A'
                ws.cell(i,10).value=id
wb.save('Email.xlsx')
wb.close()
server.logout()
